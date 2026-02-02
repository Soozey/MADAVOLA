import hashlib
import re
import unicodedata
from datetime import datetime, timezone
from dataclasses import dataclass
from io import BytesIO
from typing import Iterable

from openpyxl import load_workbook
from sqlalchemy.orm import Session

from app.common.errors import bad_request
from app.models.territory import Commune, District, Fokontany, Region, TerritoryVersion

REQUIRED_COLUMNS = {
    "region_code",
    "region_name",
    "district_code",
    "district_name",
    "commune_code",
    "commune_name",
    "fokontany_name",
}

OPTIONAL_COLUMNS = {
    "fokontany_code",
    "commune_mobile_money_msisdn",
    "latitude",
    "longitude",
}

ALIASES = {
    "region code": "region_code",
    "region name": "region_name",
    "district code": "district_code",
    "district name": "district_name",
    "commune code": "commune_code",
    "commune name": "commune_name",
    "fokontany code": "fokontany_code",
    "fokontany name": "fokontany_name",
    "commune_mobile_money_msisdn": "commune_mobile_money_msisdn",
    "commune mobile money msisdn": "commune_mobile_money_msisdn",
    "latitude": "latitude",
    "longitude": "longitude",
}


@dataclass(frozen=True)
class ImportCounts:
    regions: int
    districts: int
    communes: int
    fokontany: int


def import_territory_excel(
    db: Session, file_bytes: bytes, filename: str, version_tag: str
) -> ImportCounts:
    checksum = hashlib.sha256(file_bytes).hexdigest()

    if db.query(TerritoryVersion).filter_by(version_tag=version_tag).first():
        raise bad_request("version_tag_deja_utilise", {"version_tag": version_tag})

    workbook = load_workbook(BytesIO(file_bytes), read_only=True)
    sheet = workbook.active

    headers = _normalize_headers(next(sheet.iter_rows(min_row=1, max_row=1, values_only=True)))
    missing = REQUIRED_COLUMNS - set(headers.values())
    if missing:
        raise bad_request("colonnes_requises_manquantes", {"missing": sorted(missing)})

    rows = sheet.iter_rows(min_row=2, values_only=True)
    parsed = list(_parse_rows(rows, headers))
    if not parsed:
        raise bad_request("fichier_vide")

    with db.begin():
        current_active = (
            db.query(TerritoryVersion).filter_by(status="active").with_for_update().first()
        )
        version = TerritoryVersion(
            version_tag=version_tag,
            source_filename=filename,
            checksum_sha256=checksum,
            status="importing",
        )
        db.add(version)
        db.flush()

        regions_cache: dict[str, Region] = {}
        districts_cache: dict[tuple[str, str], District] = {}
        communes_cache: dict[tuple[str, str], Commune] = {}
        fokontany_seen: set[tuple[str, str, str]] = set()

        for row in parsed:
            region = _get_or_create_region(db, version.id, regions_cache, row)
            district = _get_or_create_district(db, version.id, districts_cache, row, region)
            commune = _get_or_create_commune(db, version.id, communes_cache, row, district)
            _create_fokontany(db, version.id, fokontany_seen, row, commune)

        if current_active:
            current_active.status = "archived"

        version.status = "active"
        version.activated_at = datetime.now(timezone.utc)

    return ImportCounts(
        regions=len(regions_cache),
        districts=len(districts_cache),
        communes=len(communes_cache),
        fokontany=len(fokontany_seen),
    )


def _normalize_headers(headers: Iterable[str | None]) -> dict[int, str]:
    normalized = {}
    for idx, header in enumerate(headers):
        if header is None:
            continue
        key = re.sub(r"\s+", " ", str(header)).strip().lower()
        key = ALIASES.get(key, key.replace(" ", "_"))
        normalized[idx] = key
    return normalized


def _parse_rows(rows: Iterable[tuple], headers: dict[int, str]) -> Iterable[dict[str, str]]:
    for row_index, row in enumerate(rows, start=2):
        data = {}
        for idx, key in headers.items():
            value = row[idx] if idx < len(row) else None
            if value is None:
                continue
            data[key] = _clean_text(value)
        missing = REQUIRED_COLUMNS - set(data.keys())
        if missing:
            raise bad_request("valeur_manquante", {"row": row_index, "missing": sorted(missing)})
        yield data


def _clean_text(value: object) -> str:
    text = str(value).strip()
    return re.sub(r"\s+", " ", text)


def _normalize_key(text: str) -> str:
    text = unicodedata.normalize("NFKD", text)
    text = "".join(c for c in text if not unicodedata.combining(c))
    return unicodedata.normalize("NFKC", text).casefold()


def _get_or_create_region(
    db: Session, version_id: int, cache: dict[str, Region], row: dict[str, str]
) -> Region:
    code = row["region_code"]
    name = row["region_name"]
    if code in cache:
        region = cache[code]
        if region.name != name:
            raise bad_request("region_incoherente", {"code": code, "name": name})
        return region
    region = Region(
        version_id=version_id,
        code=code,
        name=name,
        name_normalized=_normalize_key(name),
    )
    db.add(region)
    db.flush()
    cache[code] = region
    return region


def _get_or_create_district(
    db: Session,
    version_id: int,
    cache: dict[tuple[str, str], District],
    row: dict[str, str],
    region: Region,
) -> District:
    key = (row["region_code"], row["district_code"])
    name = row["district_name"]
    if key in cache:
        district = cache[key]
        if district.name != name:
            raise bad_request("district_incoherent", {"code": key[1], "name": name})
        return district
    district = District(
        version_id=version_id,
        region_id=region.id,
        code=key[1],
        name=name,
        name_normalized=_normalize_key(name),
    )
    db.add(district)
    db.flush()
    cache[key] = district
    return district


def _get_or_create_commune(
    db: Session,
    version_id: int,
    cache: dict[tuple[str, str], Commune],
    row: dict[str, str],
    district: District,
) -> Commune:
    key = (row["district_code"], row["commune_code"])
    name = row["commune_name"]
    if key in cache:
        commune = cache[key]
        if commune.name != name:
            raise bad_request("commune_incoherente", {"code": key[1], "name": name})
        return commune
    commune = Commune(
        version_id=version_id,
        district_id=district.id,
        code=key[1],
        name=name,
        name_normalized=_normalize_key(name),
        mobile_money_msisdn=row.get("commune_mobile_money_msisdn"),
        latitude=row.get("latitude"),
        longitude=row.get("longitude"),
    )
    db.add(commune)
    db.flush()
    cache[key] = commune
    return commune


def _create_fokontany(
    db: Session,
    version_id: int,
    seen: set[tuple[str, str, str]],
    row: dict[str, str],
    commune: Commune,
) -> None:
    code = row.get("fokontany_code") or ""
    name = row["fokontany_name"]
    name_norm = _normalize_key(name)
    key = (commune.code, code, name_norm)
    if key in seen:
        return
    if not code:
        code = None
    fokontany = Fokontany(
        version_id=version_id,
        commune_id=commune.id,
        code=code,
        name=name,
        name_normalized=name_norm,
    )
    db.add(fokontany)
    db.flush()
    seen.add(key)
